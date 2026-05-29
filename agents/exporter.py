"""
Agent 5: Lead Exporter
Genera Excel con links LinkedIn clickeables + CSV plantilla HubSpot
"""

import asyncio
from datetime import datetime
from pathlib import Path
from typing import Callable, Optional

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter


class LeadExporter:
    """Exporta leads a Excel + CSV listos para HubSpot"""

    def __init__(self, emit_fn: Optional[Callable] = None, output_dir: Path = Path("outputs")):
        self.emit = emit_fn or (lambda x: None)
        self.output_dir = output_dir
        self.output_dir.mkdir(exist_ok=True)

    async def _emit_log(self, msg: str):
        await self.emit({
            "type": "agent_log",
            "agent": "exporter",
            "level": "info",
            "message": msg,
            "timestamp": datetime.now().isoformat()
        })

    async def export(self, companies: list[dict], contacts: list[dict], params: dict) -> dict:
        """Genera Excel + CSV y retorna paths"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_file = self.output_dir / f"leads_MNR_{timestamp}.xlsx"
        csv_file = self.output_dir / f"hubspot_import_{timestamp}.csv"

        await self._emit_log(f"📊 Generando Excel con {len(contacts)} contactos...")
        await asyncio.to_thread(self._create_excel, excel_file, companies, contacts, params)

        await self._emit_log(f"📋 Generando CSV HubSpot...")
        await asyncio.to_thread(self._create_hubspot_csv, csv_file, companies, contacts)

        await self._emit_log(f"✅ Archivos generados en outputs/")

        return {
            "excel_file": excel_file.name,
            "csv_file": csv_file.name,
            "excel_path": str(excel_file),
            "csv_path": str(csv_file),
            "total_companies": len(companies),
            "total_contacts": len(contacts),
            "timestamp": timestamp
        }

    def _create_excel(self, path: Path, companies: list[dict], contacts: list[dict], params: dict):
        """Crea archivo Excel profesional con 3 hojas"""
        wb = openpyxl.Workbook()

        # ── Hoja 1: Contactos (principal) ─────────────────────────────────
        self._create_contacts_sheet(wb, contacts)

        # ── Hoja 2: Empresas ──────────────────────────────────────────────
        self._create_companies_sheet(wb, companies)

        # ── Hoja 3: Resumen Dashboard ─────────────────────────────────────
        self._create_summary_sheet(wb, companies, contacts, params)

        wb.save(str(path))

    def _header_style(self, ws, headers: list[str], col_widths: list[int]):
        """Aplica estilo a la fila de encabezado"""
        BLUE = "1F4E78"
        WHITE = "FFFFFF"

        for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(name="Calibri", bold=True, color=WHITE, size=11)
            cell.fill = PatternFill("solid", fgColor=BLUE)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(
                bottom=Side(style="medium", color=WHITE),
                right=Side(style="thin", color="AAAAAA")
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        ws.row_dimensions[1].height = 35

    def _create_contacts_sheet(self, wb: openpyxl.Workbook, contacts: list[dict]):
        """Hoja principal de contactos con links LinkedIn clickeables"""
        ws = wb.active
        ws.title = "📋 Contactos Decisores"
        ws.freeze_panes = "A2"

        headers = [
            "Empresa", "NIT", "Ciudad",
            "Nombre Contacto", "Cargo", "LinkedIn",
            "Email", "Teléfono",
            "Fuente", "Fecha"
        ]
        col_widths = [30, 16, 14, 28, 30, 50, 30, 16, 18, 14]
        self._header_style(ws, headers, col_widths)

        # Colores alternados
        LIGHT_BLUE = "EBF3FB"
        WHITE = "FFFFFF"
        GREEN = "E2EFDA"

        for row_idx, contact in enumerate(contacts, 2):
            row_color = LIGHT_BLUE if row_idx % 2 == 0 else WHITE

            data = [
                contact.get("company_name", ""),
                contact.get("company_nit", ""),
                contact.get("company_city", "") or contact.get("city", ""),
                contact.get("name", ""),
                contact.get("title", ""),
                "",  # LinkedIn (se agrega como hyperlink)
                contact.get("email", ""),
                contact.get("phone", ""),
                contact.get("source", ""),
                datetime.now().strftime("%Y-%m-%d")
            ]

            for col_idx, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.fill = PatternFill("solid", fgColor=row_color)
                cell.alignment = Alignment(vertical="center", wrap_text=False)
                cell.font = Font(name="Calibri", size=10)

            # LinkedIn URL como hyperlink
            linkedin_url = contact.get("linkedin_url", "")
            if linkedin_url:
                cell = ws.cell(row=row_idx, column=6)
                cell.value = "Ver perfil →"
                cell.hyperlink = linkedin_url
                cell.font = Font(name="Calibri", size=10, color="0563C1", underline="single")
                cell.fill = PatternFill("solid", fgColor=row_color)

            ws.row_dimensions[row_idx].height = 18

        # Añadir tabla Excel nativa
        if contacts:
            from openpyxl.worksheet.table import Table, TableStyleInfo
            table_ref = f"A1:{get_column_letter(len(headers))}{len(contacts)+1}"
            table = Table(displayName="Contactos", ref=table_ref)
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )
            ws.add_table(table)

    def _create_companies_sheet(self, wb: openpyxl.Workbook, companies: list[dict]):
        """Hoja de empresas con datos RUES"""
        ws = wb.create_sheet("🏭 Empresas RUES")
        ws.freeze_panes = "A2"

        headers = [
            "Razón Social", "NIT", "Estado RUES", "Ciudad",
            "CIIU", "Fuente", "Verificado"
        ]
        col_widths = [35, 16, 14, 16, 10, 18, 16]
        self._header_style(ws, headers, col_widths)

        for row_idx, company in enumerate(companies, 2):
            status = company.get("rues_status", "PENDIENTE")
            row_color = "E2EFDA" if status == "VIGENTE" else "FFF2CC"

            data = [
                company.get("rues_name") or company.get("name", ""),
                company.get("nit", ""),
                status,
                company.get("city", ""),
                company.get("ciiu_rues") or company.get("ciiu", ""),
                company.get("source", ""),
                company.get("rues_checked", "")[:10] if company.get("rues_checked") else ""
            ]

            for col_idx, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.fill = PatternFill("solid", fgColor=row_color)
                cell.alignment = Alignment(vertical="center")
                cell.font = Font(name="Calibri", size=10,
                                 bold=(col_idx == 3),
                                 color=("375623" if status == "VIGENTE" else "7F6000"))

    def _create_summary_sheet(self, wb, companies, contacts, params):
        """Hoja de resumen ejecutivo"""
        ws = wb.create_sheet("📊 Resumen")

        DARK_BLUE = "1F4E78"
        ORANGE = "ED7D31"

        # Título
        ws.merge_cells("A1:E1")
        ws["A1"] = "🚀 Lead GenAI · Mercado No Regulado Colombia"
        ws["A1"].font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
        ws["A1"].fill = PatternFill("solid", fgColor=DARK_BLUE)
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 40

        # KPIs
        kpis = [
            ("Empresas Encontradas", len(companies)),
            ("Vigentes en RUES", sum(1 for c in companies if c.get("rues_status") == "VIGENTE")),
            ("Decision Makers", len(contacts)),
            ("Con LinkedIn URL", sum(1 for c in contacts if c.get("linkedin_url"))),
            ("Con Email", sum(1 for c in contacts if c.get("email"))),
        ]

        for row_idx, (label, value) in enumerate(kpis, 3):
            ws.cell(row=row_idx, column=1, value=label).font = Font(name="Calibri", size=12, bold=True)
            val_cell = ws.cell(row=row_idx, column=2, value=value)
            val_cell.font = Font(name="Calibri", size=14, bold=True, color=ORANGE)
            ws.row_dimensions[row_idx].height = 25

        ws["A9"] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws["A9"].font = Font(name="Calibri", size=10, color="888888", italic=True)

        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 16

    def _create_hubspot_csv(self, path: Path, companies: list[dict], contacts: list[dict]):
        """CSV con formato de importación directa a HubSpot"""
        import csv

        # Campos HubSpot CRM
        hubspot_fields = [
            "First Name", "Last Name", "Email", "Phone",
            "Company Name", "Company NIT", "Job Title",
            "LinkedIn Profile", "City", "Country",
            "Lead Source", "Lead Status", "Notes", "Create Date"
        ]

        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=hubspot_fields)
            writer.writeheader()

            for contact in contacts:
                # Separar nombre
                full_name = contact.get("name", "").strip()
                parts = full_name.split(" ", 2)
                first_name = parts[0] if parts else ""
                last_name = " ".join(parts[1:]) if len(parts) > 1 else ""

                writer.writerow({
                    "First Name": first_name,
                    "Last Name": last_name,
                    "Email": contact.get("email", ""),
                    "Phone": contact.get("phone", ""),
                    "Company Name": contact.get("company_name", ""),
                    "Company NIT": contact.get("company_nit", ""),
                    "Job Title": contact.get("title", ""),
                    "LinkedIn Profile": contact.get("linkedin_url", ""),
                    "City": contact.get("company_city", "") or contact.get("city", ""),
                    "Country": "Colombia",
                    "Lead Source": "Lead GenAI",
                    "Lead Status": "New",
                    "Notes": f"MNR - {contact.get('source', '')}",
                    "Create Date": datetime.now().strftime("%Y-%m-%d")
                })
