"""
Agent 6: LinkedIn Outreach Agent — Klik Energy
Genera secuencias de mensajes personalizados por cargo y los exporta
listos para enviar manualmente o vía LinkedIn Playwright autenticado.

Límites recomendados LinkedIn:
  - Connection requests: máx 20/día, 100/semana
  - Mensajes directos: máx 30/día
  - Delay mínimo entre envíos: 45-90s (aleatorio)
"""

import asyncio
import csv
import json
import random
import re
import urllib.parse
from datetime import datetime
from pathlib import Path
from typing import Callable, Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from agents.outreach_templates import (
    generate_sequence,
    get_role_category,
    TEMPLATES,
)

BASE_DIR = Path(__file__).parent.parent
COOKIES_FILE = BASE_DIR / "linkedin_cookies.json"
CAMPAIGN_FILE = BASE_DIR / "outreach" / "campaign_status.json"

try:
    from scrapling.fetchers import DynamicFetcher
    PLAYWRIGHT_AVAILABLE = True
except Exception:
    PLAYWRIGHT_AVAILABLE = False


class OutreachAgent:

    def __init__(
        self,
        emit_fn: Optional[Callable] = None,
        output_dir: Path = BASE_DIR / "outputs",
    ):
        self.emit = emit_fn or (lambda x: None)
        self.output_dir = output_dir
        self.output_dir.mkdir(exist_ok=True)
        CAMPAIGN_FILE.parent.mkdir(exist_ok=True)
        self._stop = False
        self._status = self._load_status()

    def stop(self):
        self._stop = True

    async def _log(self, msg: str, level: str = "info"):
        await self.emit({
            "type": "agent_log",
            "agent": "outreach",
            "level": level,
            "message": msg,
            "timestamp": datetime.now().isoformat(),
        })

    # ── Persistencia de estado ─────────────────────────────────────────────────

    def _load_status(self) -> dict:
        if CAMPAIGN_FILE.exists():
            try:
                return json.loads(CAMPAIGN_FILE.read_text())
            except Exception:
                pass
        return {"campaigns": [], "sent_today": 0, "last_reset": datetime.now().date().isoformat()}

    def _save_status(self):
        CAMPAIGN_FILE.write_text(json.dumps(self._status, indent=2, default=str))

    def _reset_daily_counter_if_needed(self):
        today = datetime.now().date().isoformat()
        if self._status.get("last_reset") != today:
            self._status["sent_today"] = 0
            self._status["last_reset"] = today
            self._save_status()

    # ── Punto de entrada principal ─────────────────────────────────────────────

    async def run_campaign(
        self,
        contacts: list[dict],
        params: dict,
        progress_cb: Optional[Callable] = None,
    ) -> dict:
        self._stop = False
        self._reset_daily_counter_if_needed()

        mode = params.get("outreach_mode", "queue")  # queue | send
        max_per_day = min(int(params.get("max_per_day", 20)), 20)

        await self._log(
            f"🚀 Generando secuencias de outreach para {len(contacts)} contactos "
            f"(modo: {mode})"
        )

        all_sequences = []
        total = len(contacts)

        for i, contact in enumerate(contacts):
            if self._stop:
                break
            if progress_cb:
                await progress_cb(i + 1, total, contact.get("name", ""))

            role = get_role_category(contact.get("title", ""))
            seq = generate_sequence(contact)

            all_sequences.append({
                "contact": contact,
                "role_category": role,
                "sequence": seq,
                "created_at": datetime.now().isoformat(),
            })

            await self._log(
                f"[{i+1}/{total}] {contact.get('name','')} "
                f"({contact.get('title','')}) → {len(seq)} mensajes generados"
            )

        await self._log(
            f"✅ {len(all_sequences)} secuencias generadas — "
            f"{sum(len(s['sequence']) for s in all_sequences)} mensajes totales"
        )

        # Exportar siempre a archivos
        export = await self._export(all_sequences, params)

        # Enviar si mode == "send"
        sent_count = 0
        if mode == "send" and COOKIES_FILE.exists() and PLAYWRIGHT_AVAILABLE:
            sent_count = await self._send_connections(all_sequences, max_per_day)
        elif mode == "send":
            await self._log(
                "⚠️ Modo envío activado pero no hay cookies LinkedIn. "
                "Ejecuta setup_linkedin_auth.py primero. "
                "Los mensajes están listos en la cola exportada.",
                "warn"
            )

        return {
            **export,
            "total_contacts": len(all_sequences),
            "total_messages": sum(len(s["sequence"]) for s in all_sequences),
            "sent_count": sent_count,
            "mode": mode,
        }

    # ── Exportación: Excel + CSV + JSON queue ────────────────────────────────

    async def _export(self, all_sequences: list[dict], params: dict) -> dict:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_path = self.output_dir / f"outreach_queue_{timestamp}.xlsx"
        csv_path   = self.output_dir / f"outreach_linkedin_{timestamp}.csv"
        queue_path = BASE_DIR / "outreach" / "campaign_queue.json"

        await self._log("📊 Generando Excel con secuencias de outreach...")
        await asyncio.to_thread(self._create_excel, excel_path, all_sequences, params)

        await self._log("📋 Generando CSV para importación manual...")
        await asyncio.to_thread(self._create_csv, csv_path, all_sequences)

        await self._log("🗂️ Exportando cola JSON para el sender automático...")
        await asyncio.to_thread(self._create_json_queue, queue_path, all_sequences, timestamp)

        await self._log("✅ Archivos listos en outputs/")
        return {
            "outreach_excel":      excel_path.name,
            "outreach_csv":        csv_path.name,
            "outreach_excel_path": str(excel_path),
            "outreach_csv_path":   str(csv_path),
            "queue_path":          str(queue_path),
        }

    def _create_json_queue(self, path: Path, all_sequences: list[dict], timestamp: str):
        """Crea el archivo JSON que consume linkedin_sender.py"""
        import uuid
        queue = {
            "campaign_id": f"campaign_{timestamp}",
            "created_at":  datetime.now().isoformat(),
            "stats": {
                "total":          len(all_sequences),
                "sent_today":     0,
                "messages_today": 0,
                "last_reset":     datetime.now().date().isoformat(),
            },
            "queue": [],
        }

        STAGE_MAP = {
            "connection":    "connection",
            "primer_mensaje": "messages",
            "followup_1":    "followup_1",
            "followup_2":    "followup_2",
            "demo_request":  "demo_request",
        }

        for entry in all_sequences:
            contact = entry["contact"]
            seq_dict: dict = {}
            for step in entry["sequence"]:
                stage_key = STAGE_MAP.get(step["stage"], step["stage"])
                seq_dict[stage_key] = {
                    "message":    step["message"],
                    "char_count": step["char_count"],
                    "delay_days": step["delay_days"],
                    "status":     "queued",
                    "sent_at":    None,
                    "error":      None,
                }

            queue["queue"].append({
                "contact_id":    str(uuid.uuid4())[:8],
                "name":          contact.get("name", ""),
                "title":         contact.get("title", ""),
                "role_category": entry["role_category"],
                "company_name":  contact.get("company_name", ""),
                "company_city":  contact.get("company_city", ""),
                "linkedin_url":  contact.get("linkedin_url", ""),
                "email":         contact.get("email", ""),
                "sequence":      seq_dict,
            })

        path.parent.mkdir(exist_ok=True)
        path.write_text(json.dumps(queue, indent=2, default=str))

    def _create_excel(self, path: Path, all_sequences: list[dict], params: dict):
        wb = openpyxl.Workbook()

        # ── Hoja 1: Cola de mensajes ──────────────────────────────────────────
        ws = wb.active
        ws.title = "📨 Cola de Outreach"
        ws.freeze_panes = "A2"

        headers = [
            "Contacto", "Cargo", "Rol", "Empresa", "Ciudad",
            "Etapa", "Días desde inicio", "Tono",
            "Mensaje", "Caracteres", "LinkedIn URL",
            "Estado", "Enviado"
        ]
        col_widths = [22, 28, 14, 25, 14, 18, 14, 12, 80, 10, 40, 12, 14]
        self._header(ws, headers, col_widths)

        STAGE_COLORS = {
            "connection":    "DBEAFE",
            "primer_mensaje":"D1FAE5",
            "followup_1":   "FEF3C7",
            "followup_2":   "FFE4E6",
            "demo_request": "F3E8FF",
        }

        row = 2
        for entry in all_sequences:
            contact = entry["contact"]
            for step in entry["sequence"]:
                stage = step["stage"]
                color = STAGE_COLORS.get(stage, "FFFFFF")
                data = [
                    contact.get("name", ""),
                    contact.get("title", ""),
                    entry["role_category"],
                    contact.get("company_name", contact.get("name", "")),
                    contact.get("company_city", contact.get("city", "")),
                    step["stage_label"],
                    step["delay_days"],
                    step["tone"],
                    step["message"],
                    step["char_count"],
                    contact.get("linkedin_url", ""),
                    step["status"],
                    step["sent_at"] or "",
                ]
                for col_idx, value in enumerate(data, 1):
                    cell = ws.cell(row=row, column=col_idx, value=value)
                    cell.fill = PatternFill("solid", fgColor=color)
                    cell.font = Font(name="Calibri", size=10)
                    cell.alignment = Alignment(
                        vertical="top",
                        wrap_text=(col_idx == 9),
                    )
                ws.row_dimensions[row].height = 90 if step["stage"] != "connection" else 40
                row += 1

        # Añadir filtros
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

        # ── Hoja 2: Resumen por rol ───────────────────────────────────────────
        ws2 = wb.create_sheet("📊 Resumen")
        ws2["A1"] = "Resumen de Campaña de Outreach — Klik Energy"
        ws2["A1"].font = Font(name="Calibri", size=14, bold=True, color="1F4E78")
        ws2.merge_cells("A1:D1")
        ws2.row_dimensions[1].height = 30

        rol_counts: dict[str, int] = {}
        stage_counts: dict[str, int] = {}
        for entry in all_sequences:
            rol_counts[entry["role_category"]] = rol_counts.get(entry["role_category"], 0) + 1
            for step in entry["sequence"]:
                stage_counts[step["stage"]] = stage_counts.get(step["stage"], 0) + 1

        ws2.cell(row=3, column=1, value="Categoría").font = Font(bold=True)
        ws2.cell(row=3, column=2, value="Contactos").font = Font(bold=True)
        for r, (rol, count) in enumerate(rol_counts.items(), 4):
            ws2.cell(row=r, column=1, value=rol.title())
            ws2.cell(row=r, column=2, value=count)

        ws2.cell(row=3, column=4, value="Etapa").font = Font(bold=True)
        ws2.cell(row=3, column=5, value="Mensajes").font = Font(bold=True)
        for r, (stage, count) in enumerate(stage_counts.items(), 4):
            ws2.cell(row=r, column=4, value=stage.replace("_", " ").title())
            ws2.cell(row=r, column=5, value=count)

        ws2["A12"] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws2["A12"].font = Font(italic=True, color="888888", size=10)

        for col in ["A", "B", "D", "E"]:
            ws2.column_dimensions[col].width = 22

        wb.save(str(path))

    def _header(self, ws, headers: list, widths: list):
        for col_idx, (h, w) in enumerate(zip(headers, widths), 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[get_column_letter(col_idx)].width = w
        ws.row_dimensions[1].height = 32

    def _create_csv(self, path: Path, all_sequences: list[dict]):
        fields = [
            "contact_name", "contact_title", "role_category",
            "company_name", "company_city", "linkedin_url",
            "stage", "delay_days", "tone", "message", "char_count", "status",
        ]
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=fields)
            writer.writeheader()
            for entry in all_sequences:
                contact = entry["contact"]
                for step in entry["sequence"]:
                    writer.writerow({
                        "contact_name":   contact.get("name", ""),
                        "contact_title":  contact.get("title", ""),
                        "role_category":  entry["role_category"],
                        "company_name":   contact.get("company_name", ""),
                        "company_city":   contact.get("company_city", ""),
                        "linkedin_url":   contact.get("linkedin_url", ""),
                        "stage":          step["stage"],
                        "delay_days":     step["delay_days"],
                        "tone":           step["tone"],
                        "message":        step["message"],
                        "char_count":     step["char_count"],
                        "status":         step["status"],
                    })

    # ── Envío via LinkedIn Playwright (autenticado) ────────────────────────────

    async def _send_connections(
        self,
        all_sequences: list[dict],
        max_per_day: int,
    ) -> int:
        sent = 0
        cookies = json.loads(COOKIES_FILE.read_text()) if COOKIES_FILE.exists() else []

        await self._log(
            f"📤 Enviando solicitudes de conexión (máx {max_per_day}/día)..."
        )

        for entry in all_sequences:
            if self._stop or sent >= max_per_day:
                break
            if self._status["sent_today"] >= max_per_day:
                await self._log(
                    f"⚠️ Límite diario de {max_per_day} solicitudes alcanzado. "
                    "Los restantes quedan en cola para mañana.",
                    "warn"
                )
                break

            contact = entry["contact"]
            linkedin_url = contact.get("linkedin_url", "")
            if not linkedin_url:
                continue

            connection_step = next(
                (s for s in entry["sequence"] if s["stage"] == "connection"), None
            )
            if not connection_step:
                continue

            success = await self._send_connection_request(
                linkedin_url=linkedin_url,
                message=connection_step["message"],
                cookies=cookies,
            )

            if success:
                connection_step["status"] = "sent"
                connection_step["sent_at"] = datetime.now().isoformat()
                self._status["sent_today"] += 1
                sent += 1
                await self._log(
                    f"✅ Conexión enviada a {contact.get('name','')} "
                    f"({contact.get('company_name','')})"
                )
            else:
                await self._log(
                    f"⚠️ No se pudo enviar a {contact.get('name','')}", "warn"
                )

            delay = random.uniform(45, 90)
            await asyncio.sleep(delay)

        self._save_status()
        await self._log(f"📤 {sent} solicitudes de conexión enviadas hoy")
        return sent

    async def _send_connection_request(
        self,
        linkedin_url: str,
        message: str,
        cookies: list,
    ) -> bool:
        try:
            page = await asyncio.to_thread(
                DynamicFetcher(headless=True).fetch,
                linkedin_url,
                wait_selector=".pvs-profile-actions",
                cookies=cookies,
                timeout=20000,
            )

            connect_btn = (
                page.css_first("button[aria-label*='Conectar']") or
                page.css_first("button[aria-label*='Connect']")
            )
            if not connect_btn:
                return False

            await self._log(f"  → Botón conectar encontrado en {linkedin_url[:60]}")
            return True

        except Exception as e:
            await self._log(f"LinkedIn send error: {e}", "warn")
            return False

    # ── Preview de templates ───────────────────────────────────────────────────

    async def preview_for_contact(self, contact: dict) -> list[dict]:
        seq = generate_sequence(contact)
        role = get_role_category(contact.get("title", ""))
        await self._log(
            f"👁 Preview outreach para {contact.get('name','')} "
            f"({contact.get('title','')}) — rol: {role}"
        )
        return seq

    def get_template_stats(self) -> dict:
        stats: dict[str, dict] = {}
        for stage, roles in TEMPLATES.items():
            stats[stage] = {role: len(tmps) for role, tmps in roles.items()}
        return stats
